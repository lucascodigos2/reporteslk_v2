#!/usr/bin/env python3
"""
Seguimiento de exámenes para formaciones teleformación.

Cruza el calendario de días lectivos (INF_30) con el informe de progreso
de plataforma (progress CSV/XLSX) y genera un Excel visual para que el
profesorado vea quién está al día, a quién hay que recordar y quién
no se ha presentado.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ---------------------------------------------------------------------------
# Constantes de estado / colores
# ---------------------------------------------------------------------------

STATUS_HECHO = "HECHO"
STATUS_HECHO_TARDE = "HECHO TARDE"
STATUS_SUSPENSO = "SUSPENSO"
STATUS_RECORDAR = "RECORDAR"
STATUS_FUERA = "FUERA DE PLAZO"
STATUS_NO_ABIERTO = "NO ABIERTO"
STATUS_SIN_PLAZO = "PENDIENTE"
STATUS_NA = "—"

FILL = {
    STATUS_HECHO: PatternFill("solid", fgColor="C6EFCE"),
    STATUS_HECHO_TARDE: PatternFill("solid", fgColor="BDD7EE"),
    STATUS_SUSPENSO: PatternFill("solid", fgColor="C9A0DC"),
    STATUS_RECORDAR: PatternFill("solid", fgColor="FFC000"),
    STATUS_FUERA: PatternFill("solid", fgColor="FF6B6B"),
    STATUS_NO_ABIERTO: PatternFill("solid", fgColor="D9D9D9"),
    STATUS_SIN_PLAZO: PatternFill("solid", fgColor="FFF2CC"),
    "header_abierto": PatternFill("solid", fgColor="F4B183"),
    "header": PatternFill("solid", fgColor="1F4E79"),
    "kpi": PatternFill("solid", fgColor="D6EAF8"),
    "titulo": PatternFill("solid", fgColor="1F4E79"),
    "al_dia": PatternFill("solid", fgColor="C6EFCE"),
    "recordar_global": PatternFill("solid", fgColor="FFC000"),
    "retrasado": PatternFill("solid", fgColor="FF6B6B"),
    "sin_empezar": PatternFill("solid", fgColor="F4B183"),
    "leyenda": PatternFill("solid", fgColor="F2F2F2"),
}

FONT_WHITE = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
FONT_BOLD = Font(bold=True, name="Calibri", size=11)
FONT_TITLE = Font(bold=True, color="FFFFFF", name="Calibri", size=16)
FONT_NORMAL = Font(name="Calibri", size=10)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

DONE_PATTERN = re.compile(
    r"^\s*finalizado\b",
    re.IGNORECASE,
)
NOT_DONE_PATTERN = re.compile(r"no\s+finalizado", re.IGNORECASE)
# "Finalizado (no ha alcanzado la calificación de aprobado)" → presentado pero suspenso
NOT_PASSED_PATTERN = re.compile(r"no[n]?\s+.{0,8}?alcanz", re.IGNORECASE)
EQUALITY_MF_PREFIX = "FCO"


# ---------------------------------------------------------------------------
# Modelos
# ---------------------------------------------------------------------------

@dataclass
class ExamDef:
    key: str
    name: str
    pub_date: date | None
    deadline: date | None
    cod_mf: str | None
    is_final: bool
    is_equality: bool
    has_calendar: bool


@dataclass
class StudentExam:
    status: str
    completed_at: datetime | None
    raw_status: str


# ---------------------------------------------------------------------------
# Lectura de inputs
# ---------------------------------------------------------------------------

def _find_dia_column(columns: Iterable[str]) -> str:
    for col in columns:
        normalized = col.encode("utf-8", "ignore").decode("utf-8")
        if "dia" in normalized.lower().replace("í", "i") and "ano" not in normalized.lower():
            if col not in ("Data Inicio", "Data Fin", "Dia do ano"):
                return col
    # Fallback: columna índice 10 del INF_30
    cols = list(columns)
    return cols[10]


def load_calendar(path: Path) -> tuple[pd.DataFrame, list[date]]:
    df = pd.read_excel(path, header=0)
    dia_col = _find_dia_column(df.columns)
    df = df.rename(columns={dia_col: "Dia"})
    df["Dia"] = pd.to_datetime(df["Dia"], errors="coerce")
    base = df[df["Tipo"].astype(str).str.upper().str.strip() == "BASE"].copy()
    lective_days = sorted({d.date() for d in base["Dia"].dropna()})
    return base, lective_days


def nth_lective_day(start: date, lective_days: list[date], n: int = 5) -> date:
    """El día de publicación cuenta como día 1; el plazo acaba el n-ésimo lectivo."""
    window = [d for d in lective_days if d >= start]
    if not window:
        return start
    if len(window) < n:
        return window[-1]
    return window[n - 1]


def build_exam_schedule(base: pd.DataFrame, lective_days: list[date]) -> list[ExamDef]:
    exams = base[base["Examen"].astype(str).str.upper().str.strip() == "SI"].copy()
    exams = exams.drop_duplicates(subset=["Dia"]).sort_values("Dia")

    specialty: list[ExamDef] = []
    equality: list[ExamDef] = []

    for _, row in exams.iterrows():
        pub = row["Dia"].date()
        cod_mf = str(row["Cod. MF"]).strip()
        is_final = str(row["Proba Final"]).upper().strip() in {"SI", "SÍ", "YES"}
        is_equality = cod_mf.upper().startswith(EQUALITY_MF_PREFIX)
        deadline = nth_lective_day(pub, lective_days, 5)
        item = ExamDef(
            key="",
            name="",
            pub_date=pub,
            deadline=deadline,
            cod_mf=cod_mf,
            is_final=is_final,
            is_equality=is_equality,
            has_calendar=True,
        )
        (equality if is_equality else specialty).append(item)

    # Nombrar especialidad: EXAME 1..N + EXAME FINAL
    non_final = [e for e in specialty if not e.is_final]
    finals = [e for e in specialty if e.is_final]
    named: list[ExamDef] = []
    for i, exam in enumerate(non_final, start=1):
        exam.key = f"EXAME_{i}"
        exam.name = f"EXAME {i}"
        named.append(exam)
    for i, exam in enumerate(finals):
        exam.key = "EXAME_FINAL" if i == 0 else f"EXAME_FINAL_{i+1}"
        exam.name = "EXAME FINAL" if i == 0 else f"EXAME FINAL {i+1}"
        named.append(exam)

    # Igualdad: Test 1, Test final
    for i, exam in enumerate(equality):
        if exam.is_final or i == len(equality) - 1:
            exam.key = "TEST_FINAL_IGUALDAD"
            exam.name = "Test final (igualdad)"
        else:
            exam.key = f"TEST_{i+1}_IGUALDAD"
            exam.name = f"Test {i+1} (igualdad)"
        named.append(exam)

    return named


def _read_progress_raw(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        # Export LMS suele ser UTF-16 TSV
        for enc in ("utf-16", "utf-16-le", "utf-8-sig", "latin-1"):
            try:
                df = pd.read_csv(path, encoding=enc, sep="\t", header=None, dtype=str)
                if df.shape[1] > 3:
                    return df
            except Exception:
                continue
        df = pd.read_csv(path, encoding="utf-8", sep=";", header=None, dtype=str)
        return df
    return pd.read_excel(path, header=None, dtype=str)


def _is_exam_header_row(values: list[str]) -> bool:
    joined = " ".join(v for v in values if v).upper()
    return "EXAME" in joined or "PROBA" in joined or "TEST" in joined


def parse_progress(path: Path) -> tuple[list[str], pd.DataFrame]:
    """Devuelve (nombres de actividades, dataframe alumnos)."""
    raw = _read_progress_raw(path).fillna("")
    # Detectar fila de cabecera de exámenes
    header_idx = 0
    for i in range(min(3, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        if _is_exam_header_row(vals):
            header_idx = i
            break

    header = [str(v).strip() for v in raw.iloc[header_idx].tolist()]
    # Columnas de actividad: pares (estado, fecha) a partir de col 2
    activity_cols: list[tuple[int, str]] = []
    for idx, name in enumerate(header):
        if idx < 2:
            continue
        if name and not name.lower().startswith("unnamed"):
            activity_cols.append((idx, name))

    activity_names = [name for _, name in activity_cols]
    rows = []
    for i in range(header_idx + 1, len(raw)):
        row = raw.iloc[i]
        name = str(row.iloc[0]).strip()
        email = str(row.iloc[1]).strip() if len(row) > 1 else ""
        if not name or name.lower() in {"nan", "none"}:
            continue
        # Filtrar filas basura / totales
        if name.upper() in {"ESTADO GLOBAL", "CALIFICACION ESPECIALIDAD", "CALIFICACIÓN MT"}:
            continue
        record = {"alumno": name, "email": email}
        for col_idx, act_name in activity_cols:
            status = str(row.iloc[col_idx]).strip() if col_idx < len(row) else ""
            date_val = ""
            if col_idx + 1 < len(row):
                next_val = str(row.iloc[col_idx + 1]).strip()
                # Si la siguiente columna no es otro examen, es la fecha
                if next_val and next_val not in activity_names:
                    date_val = next_val
            record[f"status::{act_name}"] = status
            record[f"date::{act_name}"] = date_val
        rows.append(record)

    return activity_names, pd.DataFrame(rows)


def match_activities_to_schedule(
    activity_names: list[str], schedule: list[ExamDef]
) -> list[tuple[str, ExamDef | None]]:
    """Asocia cada actividad del progress a su definición de calendario."""
    specialty_non_final = [e for e in schedule if not e.is_equality and not e.is_final]
    specialty_final = [e for e in schedule if not e.is_equality and e.is_final]
    equality = [e for e in schedule if e.is_equality]

    matched: list[tuple[str, ExamDef | None]] = []
    exam_counter = 0
    equality_counter = 0

    for name in activity_names:
        upper = name.upper()
        exam_def: ExamDef | None = None

        if "PROBA INICIAL" in upper or "PRUEBA INICIAL" in upper:
            exam_def = ExamDef(
                key="PROBA_INICIAL",
                name=name,
                pub_date=None,
                deadline=None,
                cod_mf=None,
                is_final=False,
                is_equality=False,
                has_calendar=False,
            )
        elif "IGUALDAD" in upper or "FCO" in upper or upper.startswith("TEST"):
            if equality_counter < len(equality):
                exam_def = equality[equality_counter]
                exam_def.name = name
                equality_counter += 1
            else:
                exam_def = ExamDef(
                    key=f"TEST_EXTRA_{equality_counter}",
                    name=name,
                    pub_date=None,
                    deadline=None,
                    cod_mf="FCO",
                    is_final="FINAL" in upper,
                    is_equality=True,
                    has_calendar=False,
                )
                equality_counter += 1
        elif "FINAL" in upper and "TEST" not in upper:
            if specialty_final:
                exam_def = specialty_final[0]
                exam_def.name = name
            else:
                exam_def = ExamDef(
                    key="EXAME_FINAL",
                    name=name,
                    pub_date=None,
                    deadline=None,
                    cod_mf=None,
                    is_final=True,
                    is_equality=False,
                    has_calendar=False,
                )
        elif "EXAME" in upper or "EXAMEN" in upper:
            if exam_counter < len(specialty_non_final):
                exam_def = specialty_non_final[exam_counter]
                exam_def.name = name
                exam_counter += 1
            else:
                exam_counter += 1
                exam_def = ExamDef(
                    key=f"EXAME_EXTRA_{exam_counter}",
                    name=name,
                    pub_date=None,
                    deadline=None,
                    cod_mf=None,
                    is_final=False,
                    is_equality=False,
                    has_calendar=False,
                )
        else:
            exam_def = ExamDef(
                key=re.sub(r"\W+", "_", name.upper())[:40],
                name=name,
                pub_date=None,
                deadline=None,
                cod_mf=None,
                is_final=False,
                is_equality=False,
                has_calendar=False,
            )

        matched.append((name, exam_def))
    return matched


def parse_completion_date(value: str) -> datetime | None:
    if not value or value.lower() in {"nan", "none", "nat"}:
        return None
    value = value.strip()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(value, dayfirst=True).to_pydatetime()
    except Exception:
        return None


def is_completed(raw_status: str) -> bool:
    """El alumno finalizó la actividad (aprobada o suspensa)."""
    if not raw_status:
        return False
    if NOT_DONE_PATTERN.search(raw_status):
        return False
    return bool(DONE_PATTERN.search(raw_status.strip()))


def is_passed(raw_status: str) -> bool:
    """Finalizada y aprobada (no 'no ha alcanzado la calificación de aprobado')."""
    if not is_completed(raw_status):
        return False
    return not NOT_PASSED_PATTERN.search(raw_status)


def classify_exam(
    raw_status: str,
    completed_at: datetime | None,
    exam: ExamDef,
    as_of: date,
) -> str:
    done = is_completed(raw_status)
    passed = is_passed(raw_status)

    def done_status(late: bool) -> str:
        # Presentado pero suspenso: lo hizo (no hay que recordar) pero no aprobó.
        if not passed:
            return STATUS_SUSPENSO
        return STATUS_HECHO_TARDE if late else STATUS_HECHO

    if not exam.has_calendar or exam.pub_date is None or exam.deadline is None:
        if done:
            return done_status(late=False)
        return STATUS_SIN_PLAZO if raw_status else STATUS_NA

    if as_of < exam.pub_date:
        return done_status(late=False) if done else STATUS_NO_ABIERTO

    if done:
        late = bool(completed_at and completed_at.date() > exam.deadline)
        return done_status(late)

    if exam.pub_date <= as_of <= exam.deadline:
        return STATUS_RECORDAR

    return STATUS_FUERA


def calendar_statuses(
    matrix_row: dict[str, StudentExam],
    matched: list[tuple[str, ExamDef | None]],
) -> list[str]:
    """Estados que cuentan para el estado global: solo exámenes del calendario.

    Las actividades que no aparecen en el informe de días lectivos (Examen=SI),
    como la proba inicial, son informativas y no influyen en si el alumno va al día.
    """
    return [
        matrix_row[act].status
        for act, exam in matched
        if exam and exam.has_calendar
    ]


def global_student_status(statuses: list[str]) -> str:
    """Prioridad: RETRASADO > RECORDAR > SIN EMPEZAR > AL DÍA."""
    relevant = [s for s in statuses if s not in {STATUS_NA, STATUS_NO_ABIERTO}]
    if not relevant:
        return "SIN ACTIVIDAD"
    has_overdue = STATUS_FUERA in relevant
    has_remind = STATUS_RECORDAR in relevant
    if has_overdue and has_remind:
        return "RETRASADO + RECORDAR"
    if has_overdue:
        return "RETRASADO"
    if has_remind:
        return "RECORDAR"
    done_states = {STATUS_HECHO, STATUS_HECHO_TARDE, STATUS_SUSPENSO}
    pending = [s for s in relevant if s == STATUS_SIN_PLAZO]
    done = [s for s in relevant if s in done_states]
    if pending and not done:
        return "SIN EMPEZAR"
    if pending:
        return "RECORDAR"
    return "AL DÍA"


def predict_results(
    statuses_by_key: dict[str, str],
    schedule: list[ExamDef],
) -> dict[str, str]:
    """Previsión según reglas del documento de especificaciones."""
    specialty = [e for e in schedule if e and not e.is_equality and e.has_calendar]
    equality = [e for e in schedule if e and e.is_equality and e.has_calendar]

    def done_for(exam: ExamDef) -> bool:
        return statuses_by_key.get(exam.name, STATUS_NA) in {STATUS_HECHO, STATUS_HECHO_TARDE}

    def late_or_missing(exam: ExamDef) -> bool:
        # Solo cuentan los ya vencidos / no presentados a tiempo
        return statuses_by_key.get(exam.name, STATUS_NA) == STATUS_FUERA

    n_spec = len(specialty)
    if n_spec == 0:
        return {
            "estado_global": "N/D",
            "calificacion_especialidad": "N/D",
            "calificacion_igualdad": "N/D",
        }

    done_spec = sum(1 for e in specialty if done_for(e))
    missing_late = sum(1 for e in specialty if late_or_missing(e))
    pct_done = done_spec / n_spec
    pct_missing = missing_late / n_spec

    if pct_missing >= 0.25:
        estado = "BAJA"
    elif pct_done >= 0.75:
        estado = "FINALIZADO"
    else:
        estado = "EN CURSO"

    non_final = [e for e in specialty if not e.is_final]
    finals = [e for e in specialty if e.is_final]
    if non_final:
        pct_non_final = sum(1 for e in non_final if done_for(e)) / len(non_final)
    else:
        pct_non_final = 0.0
    final_ok = all(done_for(e) for e in finals) if finals else False
    cal_esp = "APTO" if (pct_non_final >= 0.75 and final_ok) else "NO APTO"

    if equality:
        cal_eq = "APTO" if all(done_for(e) for e in equality) else "NO APTO"
    else:
        cal_eq = "N/D"

    return {
        "estado_global": estado,
        "calificacion_especialidad": cal_esp,
        "calificacion_igualdad": cal_eq,
        "pct_hechos": round(pct_done * 100, 1),
        "hechos": done_spec,
        "total_especialidad": n_spec,
    }


# ---------------------------------------------------------------------------
# Excel de salida
# ---------------------------------------------------------------------------

def _style_header(cell, fill=None):
    cell.fill = fill or FILL["header"]
    cell.font = FONT_WHITE
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = THIN


def _write_cell(ws, row, col, value, fill=None, font=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN
    cell.font = font or FONT_NORMAL
    cell.alignment = align or Alignment(vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    return cell


def lective_days_remaining(as_of: date, deadline: date, lective_days: list[date]) -> int:
    """Días lectivos que quedan hasta el límite (incluyendo hoy si es lectivo y <= límite)."""
    return sum(1 for d in lective_days if as_of <= d <= deadline)


def build_workbook(
    students: pd.DataFrame,
    matched: list[tuple[str, ExamDef | None]],
    matrix: dict[str, dict[str, StudentExam]],
    schedule: list[ExamDef],
    as_of: date,
    course_code: str,
    lective_days: list[date],
) -> Workbook:
    wb = Workbook()

    # ----- Resumen -----
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = f"Seguimiento de exámenes — {course_code}"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL["titulo"]
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Fecha de referencia: {as_of.strftime('%d/%m/%Y')}"
    ws["A2"].font = FONT_BOLD

    # Conteos globales
    globals_count = {
        "AL DÍA": 0,
        "RECORDAR": 0,
        "RETRASADO": 0,
        "RETRASADO + RECORDAR": 0,
        "SIN EMPEZAR": 0,
        "SIN ACTIVIDAD": 0,
    }
    remind_rows = []
    overdue_rows = []

    for _, stu in students.iterrows():
        alumno = stu["alumno"]
        gstat = global_student_status(calendar_statuses(matrix[alumno], matched))
        globals_count[gstat] = globals_count.get(gstat, 0) + 1
        open_pending = [
            (act, exam)
            for act, exam in matched
            if matrix[alumno][act].status == STATUS_RECORDAR
        ]
        overdue = [
            (act, exam)
            for act, exam in matched
            if matrix[alumno][act].status == STATUS_FUERA
        ]
        if open_pending:
            remind_rows.append((alumno, stu["email"], open_pending, gstat))
        if overdue:
            overdue_rows.append((alumno, stu["email"], overdue, gstat))

    ws["A4"] = "Indicadores"
    ws["A4"].font = FONT_BOLD
    n_recordar = globals_count.get("RECORDAR", 0) + globals_count.get("RETRASADO + RECORDAR", 0)
    n_retrasado = globals_count.get("RETRASADO", 0) + globals_count.get("RETRASADO + RECORDAR", 0)
    kpis = [
        ("Alumnos totales", len(students)),
        ("Al día", globals_count.get("AL DÍA", 0)),
        ("A recordar (tienen examen en plazo)", n_recordar),
        ("Retrasados / no presentados", n_retrasado),
        ("Sin empezar", globals_count.get("SIN EMPEZAR", 0)),
        ("Exámenes en plazo hoy", sum(1 for _, e in matched if e and e.has_calendar and e.pub_date and e.deadline and e.pub_date <= as_of <= e.deadline)),
    ]
    for i, (label, value) in enumerate(kpis):
        _write_cell(ws, 5 + i, 1, label, FILL["kpi"], FONT_BOLD)
        _write_cell(ws, 5 + i, 2, value, FILL["kpi"], FONT_BOLD)

    # Leyenda
    ws["A12"] = "Leyenda de estados por examen"
    ws["A12"].font = FONT_BOLD
    legend = [
        (STATUS_HECHO, "Completado y aprobado dentro de plazo"),
        (STATUS_HECHO_TARDE, "Completado y aprobado fuera de plazo"),
        (STATUS_SUSPENSO, "Presentado pero no aprobado (no cuenta como apto)"),
        (STATUS_RECORDAR, "En plazo y pendiente → contactar al alumno"),
        (STATUS_FUERA, "No se presentó / plazo vencido"),
        (STATUS_NO_ABIERTO, "Examen aún no publicado"),
        (STATUS_SIN_PLAZO, "Pendiente (sin fecha en calendario)"),
    ]
    for i, (st, desc) in enumerate(legend):
        _write_cell(ws, 13 + i, 1, st, FILL[st], FONT_BOLD)
        _write_cell(ws, 13 + i, 2, desc)

    ws["A20"] = (
        "Criterio de plazo: desde la fecha de publicación del examen se cuentan "
        "5 días lectivos del curso (el día de publicación es el 1º)."
    )
    ws.merge_cells("A20:G20")

    # Chart data
    ws["D4"] = "Distribución alumnos"
    ws["D4"].font = FONT_BOLD
    chart_labels = ["AL DÍA", "RECORDAR", "RETRASADO", "RETRASADO + RECORDAR", "SIN EMPEZAR"]
    for i, lab in enumerate(chart_labels):
        _write_cell(ws, 5 + i, 4, lab)
        _write_cell(ws, 5 + i, 5, globals_count.get(lab, 0))
    chart = BarChart()
    chart.title = "Estado global alumnos"
    chart.y_axis.title = "Alumnos"
    data = Reference(ws, min_col=5, min_row=4, max_row=9)
    cats = Reference(ws, min_col=4, min_row=5, max_row=9)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, "D10")

    for col, width in {1: 36, 2: 55, 3: 14, 4: 16, 5: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ----- Recordatorios (prioridad) -----
    wr = wb.create_sheet("Recordatorios", 1)
    wr["A1"] = "PRIORIDAD: alumnos con examen en plazo pendiente"
    wr["A1"].font = FONT_TITLE
    wr["A1"].fill = FILL["titulo"]
    wr.merge_cells("A1:F1")
    wr["A2"] = (
        "Estos alumnos todavía están a tiempo. Conviene recordarles ahora "
        "para maximizar la finalización del curso."
    )
    wr.merge_cells("A2:F2")

    headers = ["Alumno", "Email", "Examen pendiente", "Publicado", "Límite", "Días lectivos restantes"]
    for c, h in enumerate(headers, 1):
        _style_header(wr.cell(1 + 3, c, h))

    flat = []
    for alumno, email, pending, _ in remind_rows:
        for act, exam in pending:
            if not exam or not exam.deadline:
                continue
            flat.append((alumno, email, act, exam))

    row_i = 5
    flat_sorted = sorted(
        flat,
        key=lambda x: (
            lective_days_remaining(as_of, x[3].deadline, lective_days) if x[3].deadline else 999,
            x[3].deadline or date.max,
            x[0],
            x[2],
        ),
    )
    for alumno, email, act, exam in flat_sorted:
        days_left = lective_days_remaining(as_of, exam.deadline, lective_days)
        fill = FILL[STATUS_RECORDAR]
        values = [
            alumno,
            email,
            act,
            exam.pub_date.strftime("%d/%m/%Y") if exam.pub_date else "",
            exam.deadline.strftime("%d/%m/%Y") if exam.deadline else "",
            days_left,
        ]
        for c, v in enumerate(values, 1):
            _write_cell(wr, row_i, c, v, fill, FONT_BOLD if c == 1 else FONT_NORMAL)
        row_i += 1

    if row_i == 5:
        _write_cell(
            wr, 5, 1,
            "No hay alumnos con exámenes en plazo pendientes hoy.",
            font=FONT_BOLD,
        )

    wr.column_dimensions["A"].width = 32
    wr.column_dimensions["B"].width = 36
    wr.column_dimensions["C"].width = 42
    wr.column_dimensions["D"].width = 14
    wr.column_dimensions["E"].width = 14
    wr.column_dimensions["F"].width = 22

    # ----- Fuera de plazo -----
    wo = wb.create_sheet("Fuera de plazo")
    wo["A1"] = "Alumnos con exámenes fuera de plazo (no presentados)"
    wo["A1"].font = FONT_TITLE
    wo["A1"].fill = FILL["titulo"]
    wo.merge_cells("A1:E1")
    headers = ["Alumno", "Email", "Examen", "Publicado", "Límite"]
    for c, h in enumerate(headers, 1):
        _style_header(wo.cell(3, c, h))
    r = 4
    for alumno, email, overdue, _ in sorted(overdue_rows, key=lambda x: x[0]):
        for act, exam in overdue:
            vals = [
                alumno,
                email,
                act,
                exam.pub_date.strftime("%d/%m/%Y") if exam and exam.pub_date else "",
                exam.deadline.strftime("%d/%m/%Y") if exam and exam.deadline else "",
            ]
            for c, v in enumerate(vals, 1):
                _write_cell(wo, r, c, v, FILL[STATUS_FUERA])
            r += 1
    if r == 4:
        _write_cell(wo, 4, 1, "Nadie fuera de plazo.", font=FONT_BOLD)
    for col, width in {"A": 32, "B": 36, "C": 42, "D": 14, "E": 14}.items():
        wo.column_dimensions[col].width = width

    # ----- Matriz seguimiento -----
    wm = wb.create_sheet("Seguimiento", 2)
    wm["A1"] = "Matriz de seguimiento por alumno y examen"
    wm["A1"].font = FONT_TITLE
    wm["A1"].fill = FILL["titulo"]
    last_col = 3 + len(matched)
    wm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    # Fila periodos
    wm.cell(3, 1, "Periodo (pub → límite)")
    _style_header(wm.cell(3, 1))
    wm.cell(3, 2, "")
    _style_header(wm.cell(3, 2))
    wm.cell(3, 3, "")
    _style_header(wm.cell(3, 3))
    for i, (act, exam) in enumerate(matched):
        col = 4 + i
        if exam and exam.has_calendar and exam.pub_date and exam.deadline:
            text = f"{exam.pub_date.strftime('%d/%m')} → {exam.deadline.strftime('%d/%m/%Y')}"
            open_now = exam.pub_date <= as_of <= exam.deadline
        else:
            text = "Sin plazo SIFO"
            open_now = False
        cell = wm.cell(3, col, text)
        _style_header(cell, FILL["header_abierto"] if open_now else FILL["header"])

    # Cabecera nombres
    headers_m = ["Alumno", "Email", "Estado global"] + [act for act, _ in matched]
    for c, h in enumerate(headers_m, 1):
        cell = wm.cell(4, c, h)
        exam = matched[c - 4][1] if c >= 4 else None
        open_now = bool(
            exam and exam.has_calendar and exam.pub_date and exam.deadline and exam.pub_date <= as_of <= exam.deadline
        )
        _style_header(cell, FILL["header_abierto"] if open_now else FILL["header"])

    # Datos
    for r_i, (_, stu) in enumerate(students.iterrows(), start=5):
        alumno = stu["alumno"]
        email = stu["email"]
        gstat = global_student_status(calendar_statuses(matrix[alumno], matched))
        gfill = {
            "AL DÍA": FILL["al_dia"],
            "RECORDAR": FILL["recordar_global"],
            "RETRASADO": FILL["retrasado"],
            "RETRASADO + RECORDAR": FILL["retrasado"],
            "SIN EMPEZAR": FILL["sin_empezar"],
        }.get(gstat)

        _write_cell(wm, r_i, 1, alumno, font=FONT_BOLD)
        _write_cell(wm, r_i, 2, email)
        _write_cell(wm, r_i, 3, gstat, gfill, FONT_BOLD, Alignment(horizontal="center", vertical="center"))

        for c_i, (act, exam) in enumerate(matched):
            se = matrix[alumno][act]
            label = se.status
            if se.completed_at and se.status in {STATUS_HECHO, STATUS_HECHO_TARDE, STATUS_SUSPENSO}:
                label = f"{se.status}\n{se.completed_at.strftime('%d/%m %H:%M')}"
            fill = FILL.get(se.status)
            _write_cell(
                wm,
                r_i,
                4 + c_i,
                label,
                fill,
                FONT_BOLD if se.status == STATUS_RECORDAR else FONT_NORMAL,
                Alignment(horizontal="center", vertical="center", wrap_text=True),
            )

    wm.column_dimensions["A"].width = 32
    wm.column_dimensions["B"].width = 34
    wm.column_dimensions["C"].width = 14
    for i in range(len(matched)):
        wm.column_dimensions[get_column_letter(4 + i)].width = 16
    wm.row_dimensions[3].height = 30
    wm.row_dimensions[4].height = 45
    wm.freeze_panes = "D5"
    wm.auto_filter.ref = f"A4:{get_column_letter(last_col)}{4 + len(students)}"

    # ----- Calendario -----
    wc = wb.create_sheet("Calendario exámenes")
    wc["A1"] = "Calendario de exámenes (Tipo BASE + Examen = SI)"
    wc["A1"].font = FONT_TITLE
    wc["A1"].fill = FILL["titulo"]
    wc.merge_cells("A1:H1")
    heads = [
        "Actividad progress",
        "Publicación",
        "Límite (5º lectivo)",
        "Estado plazo hoy",
        "Módulo",
        "Prueba final",
        "Igualdad",
        "En calendario",
    ]
    for c, h in enumerate(heads, 1):
        _style_header(wc.cell(3, c, h))
    for r_i, (act, exam) in enumerate(matched, start=4):
        if exam and exam.has_calendar and exam.pub_date and exam.deadline:
            if as_of < exam.pub_date:
                plazo = "Aún no abierto"
                fill = FILL[STATUS_NO_ABIERTO]
            elif as_of <= exam.deadline:
                plazo = "ABIERTO HOY"
                fill = FILL[STATUS_RECORDAR]
            else:
                plazo = "Cerrado"
                fill = FILL[STATUS_FUERA]
            pub = exam.pub_date.strftime("%d/%m/%Y")
            dl = exam.deadline.strftime("%d/%m/%Y")
            mf = exam.cod_mf or ""
            is_final = "SI" if exam.is_final else "NO"
            is_eq = "SI" if exam.is_equality else "NO"
            cal = "SI"
        else:
            plazo, fill = "Sin plazo SIFO", FILL[STATUS_SIN_PLAZO]
            pub = dl = mf = is_final = is_eq = ""
            cal = "NO"
            if exam:
                is_final = "SI" if exam.is_final else "NO"
                is_eq = "SI" if exam.is_equality else "NO"
                mf = exam.cod_mf or ""
        vals = [act, pub, dl, plazo, mf, is_final, is_eq, cal]
        for c, v in enumerate(vals, 1):
            _write_cell(wc, r_i, c, v, fill if c == 4 else None, FONT_BOLD if c == 4 else FONT_NORMAL)
    for col, width in {"A": 42, "B": 14, "C": 18, "D": 16, "E": 12, "F": 12, "G": 10, "H": 14}.items():
        wc.column_dimensions[col].width = width

    # ----- Previsión -----
    wp = wb.create_sheet("Previsión")
    wp["A1"] = "Previsión de resultados (reglas de especificación)"
    wp["A1"].font = FONT_TITLE
    wp["A1"].fill = FILL["titulo"]
    wp.merge_cells("A1:H1")
    wp["A2"] = (
        "BAJA: ≥25% pruebas especialidad sin hacer/fuera de plazo. "
        "FINALIZADO: ≥75% hechas. "
        "APTO especialidad: ≥75% no-finales + prueba final. "
        "APTO igualdad: las 2 pruebas del módulo."
    )
    wp.merge_cells("A2:H2")
    heads = [
        "Alumno",
        "Email",
        "Estado global seguimiento",
        "Previsión estado",
        "% hechos especialidad",
        "Hechos / Total",
        "Calif. especialidad",
        "Calif. igualdad",
    ]
    for c, h in enumerate(heads, 1):
        _style_header(wp.cell(4, c, h))

    for r_i, (_, stu) in enumerate(students.iterrows(), start=5):
        alumno = stu["alumno"]
        statuses = {act: matrix[alumno][act].status for act, _ in matched}
        gstat = global_student_status(calendar_statuses(matrix[alumno], matched))
        pred = predict_results(statuses, [e for _, e in matched if e])
        vals = [
            alumno,
            stu["email"],
            gstat,
            pred["estado_global"],
            pred.get("pct_hechos", ""),
            f"{pred.get('hechos', '')}/{pred.get('total_especialidad', '')}",
            pred["calificacion_especialidad"],
            pred["calificacion_igualdad"],
        ]
        gfill = {
            "AL DÍA": FILL["al_dia"],
            "RECORDAR": FILL["recordar_global"],
            "RETRASADO": FILL["retrasado"],
            "RETRASADO + RECORDAR": FILL["retrasado"],
            "SIN EMPEZAR": FILL["sin_empezar"],
        }.get(gstat)
        for c, v in enumerate(vals, 1):
            fill = gfill if c == 3 else None
            if c == 4 and v == "BAJA":
                fill = FILL[STATUS_FUERA]
            elif c == 4 and v == "FINALIZADO":
                fill = FILL[STATUS_HECHO]
            _write_cell(wp, r_i, c, v, fill, FONT_BOLD if c in {1, 3, 4} else FONT_NORMAL)

    for col, width in {"A": 32, "B": 34, "C": 18, "D": 16, "E": 18, "F": 14, "G": 18, "H": 16}.items():
        wp.column_dimensions[col].width = width

    # ----- Alumnos al día -----
    wa = wb.create_sheet("Al día")
    wa["A1"] = "Alumnos al día (todos los exámenes ya abiertos están hechos)"
    wa["A1"].font = FONT_TITLE
    wa["A1"].fill = FILL["titulo"]
    wa.merge_cells("A1:C1")
    for c, h in enumerate(["Alumno", "Email", "Estado"], 1):
        _style_header(wa.cell(3, c, h))
    r = 4
    for _, stu in students.iterrows():
        alumno = stu["alumno"]
        gstat = global_student_status(calendar_statuses(matrix[alumno], matched))
        if gstat == "AL DÍA":
            for c, v in enumerate([alumno, stu["email"], gstat], 1):
                _write_cell(wa, r, c, v, FILL["al_dia"])
            r += 1
    if r == 4:
        _write_cell(wa, 4, 1, "Ningún alumno está completamente al día.")
    wa.column_dimensions["A"].width = 32
    wa.column_dimensions["B"].width = 36
    wa.column_dimensions["C"].width = 14

    return wb


# ---------------------------------------------------------------------------
# Orquestación
# ---------------------------------------------------------------------------

def run(calendar_path: Path, progress_path: Path, output_path: Path, as_of: date) -> Path:
    base, lective_days = load_calendar(calendar_path)
    schedule = build_exam_schedule(base, lective_days)
    activity_names, students = parse_progress(progress_path)
    matched = match_activities_to_schedule(activity_names, schedule)

    matrix: dict[str, dict[str, StudentExam]] = {}
    for _, stu in students.iterrows():
        alumno = stu["alumno"]
        matrix[alumno] = {}
        for act, exam in matched:
            raw = str(stu.get(f"status::{act}", "") or "")
            completed_at = parse_completion_date(str(stu.get(f"date::{act}", "") or ""))
            assert exam is not None
            status = classify_exam(raw, completed_at, exam, as_of)
            matrix[alumno][act] = StudentExam(status=status, completed_at=completed_at, raw_status=raw)

    course_code = str(base["Cod. Curso"].iloc[0]) if len(base) else calendar_path.stem
    wb = build_workbook(
        students, matched, matrix, schedule, as_of, course_code, lective_days
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    default_docs = Path(__file__).resolve().parent / "00 DOCS LK"
    parser = argparse.ArgumentParser(
        description="Genera Excel de seguimiento de exámenes para profesorado."
    )
    parser.add_argument(
        "--calendario",
        type=Path,
        default=default_docs / "INF_30_20260610_153628_1653_informe dias lectivos curso.xlsx",
        help="Excel INF_30 de días lectivos del curso",
    )
    parser.add_argument(
        "--progress",
        type=Path,
        default=default_docs
        / "progress.2026_001653_ifct0019_inteligencia_artificial_aplicada_a_la_empresa.csv",
        help="CSV/XLSX de progreso exportado de plataforma",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        default=None,
        help="Ruta del Excel de salida",
    )
    parser.add_argument(
        "--fecha",
        type=str,
        default=None,
        help="Fecha de referencia DD/MM/YYYY (por defecto: hoy)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.fecha:
        as_of = datetime.strptime(args.fecha, "%d/%m/%Y").date()
    else:
        as_of = date.today()

    if args.salida is None:
        stamp = as_of.strftime("%Y%m%d")
        args.salida = (
            Path(__file__).resolve().parent
            / "salida"
            / f"seguimiento_examenes_{stamp}.xlsx"
        )

    out = run(args.calendario, args.progress, args.salida, as_of)
    print(f"Informe generado: {out}")


if __name__ == "__main__":
    main()
